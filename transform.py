#!/usr/bin/env python3
# coding=utf-8

from docx import Document
import openpyxl

class Member(object):
    def __init__(self, name, gender, relation):
        self.name = name
        self.gender = gender
        self.relation = relation

class Family(object):
    def __init__(self, index, holderName):
        self.index = index
        self.holderName = holderName
        self.members = []

def loadFamiliesFromXlsxFile(fileName):
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook.active
    cell = sheet.cell(row = 3, column = 2)
    print cell.value
    familys = []
    for row in sheet.iter_rows(min_row = 3, min_col = 1):
        if row[0].value is not None and len(str(row[0].value).strip()) != 0:
            family = Family(row[0], row[1])
            familys.append(family)
        member = Member(row[2], row[3], row[4])
        family.members.append(member)
    print str(len(familys)) + " families are found!"
    return familys

class WordDoc(object):
    def __init__(self, doc, holderNameCell, rows):
        self.wordDoc = doc
        self.holderNameCell = holderNameCell
        self.rows = rows

def parseDocTemplate(template):
    wordDoc = Document(template)
    mCell = None
    mRows = []
    idx = -1
    table = wordDoc.tables[0]
    find2 = False
    for row in table.rows:
        find = False
        for cell in row.cells:
            if u"姓名" in cell.text:
                find2 = True
            if u"申请人" in cell.text:
                find  = True
            if find or find2:
                break
        if find:
            mCell = row.cells[3]
        if find2:
            idx = idx + 1
            if idx >= 1 and idx <= 10:
                mRows.append(row)
    mDoc = WordDoc(wordDoc, mCell, mRows)
    return mDoc

def object2string(obj):
    if obj.value is None:
        return ""
    else:
        return obj.value

def generateDocFileForOneFamily(mDoc, family):
    # clean data
    mDoc.holderNameCell.text = ''
    for idx in range(10):
        mDoc.rows[idx].cells[0].text = ''
        mDoc.rows[idx].cells[2].text = ''
        mDoc.rows[idx].cells[3].text = ''
    # fill new data
    mDoc.holderNameCell.text = object2string(family.holderName)
    for idx in range(len(family.members)):
        member = family.members[idx]
        mDoc.rows[idx].cells[0].text = object2string(member.name)
        mDoc.rows[idx].cells[2].text = object2string(member.relation)
        mDoc.rows[idx].cells[3].text = object2string(member.gender)
    fileName2 = str(family.index.value) + object2string(family.holderName) + '.docx'
    mDoc.wordDoc.save(fileName2)

fileName = "familysTable.xlsx"
template = 'template.docx'
mDoc = parseDocTemplate(template)
familys = loadFamiliesFromXlsxFile(fileName)
for family in familys:
    if family is None:
        continue
    print family.index
    generateDocFileForOneFamily(mDoc, family)